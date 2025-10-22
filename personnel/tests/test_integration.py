"""Integration tests for the Binance Tax Report Generator.

This module tests the complete end-to-end workflow of generating tax reports,
using mocked API responses with realistic test data.
"""

import unittest
import os
import tempfile
import shutil
from datetime import datetime, date
from decimal import Decimal
from unittest.mock import Mock, patch, MagicMock
from openpyxl import load_workbook

from config.config import Config
from clients.binance_client import BinanceClient
from clients.frankfurter_client import FrankfurterClient
from calculators.flat_tax_calculator import FlatTaxCalculator
from calculators.portfolio_calculator import PortfolioValueCalculator
from writers.excel_writer import ExcelReportWriter, TaxReportRow
from writers.pdf_writer import PDFReportWriter
from models import FiatOperation
from generate_tax_report import generate_tax_report


class TestIntegration(unittest.TestCase):
    """Integration tests for the complete tax report generation workflow"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.test_dir = tempfile.mkdtemp()
        self.test_year = 2025
        
        # Create test binance_keys file
        self.keys_file = os.path.join(self.test_dir, "binance_keys")
        with open(self.keys_file, 'w') as f:
            f.write("BINANCE_API_KEY='test_api_key'\n")
            f.write("BINANCE_SECRET_KEY='test_secret_key'\n")
        
        # Test data based on example Excel file
        self.test_operations = [
            {
                'date': datetime(2025, 1, 10, 12, 0, 0),
                'type': 'Dépôt',
                'amount': Decimal('10000.00'),
                'timestamp': 1736510400000,
                'portfolio_usd': Decimal('10869.57'),  # 10000 EUR / 0.92
                'exchange_rate': Decimal('0.92')
            },
            {
                'date': datetime(2025, 3, 15, 14, 30, 0),
                'type': 'Retrait',
                'amount': Decimal('3000.00'),
                'timestamp': 1742302200000,
                'portfolio_usd': Decimal('8695.65'),  # 8000 EUR / 0.92
                'exchange_rate': Decimal('0.92'),
                'portfolio_eur': Decimal('8000.00')
            },
            {
                'date': datetime(2025, 4, 5, 10, 15, 0),
                'type': 'Dépôt',
                'amount': Decimal('2000.00'),
                'timestamp': 1744027500000,
                'portfolio_usd': Decimal('13043.48'),  # 12000 EUR / 0.92
                'exchange_rate': Decimal('0.92')
            },
            {
                'date': datetime(2025, 6, 20, 16, 45, 0),
                'type': 'Retrait',
                'amount': Decimal('5000.00'),
                'timestamp': 1750611900000,
                'portfolio_usd': Decimal('16304.35'),  # 15000 EUR / 0.92
                'exchange_rate': Decimal('0.92'),
                'portfolio_eur': Decimal('15000.00')
            },
            {
                'date': datetime(2025, 9, 10, 9, 0, 0),
                'type': 'Retrait',
                'amount': Decimal('4000.00'),
                'timestamp': 1757487600000,
                'portfolio_usd': Decimal('10869.57'),  # 10000 EUR / 0.92
                'exchange_rate': Decimal('0.92'),
                'portfolio_eur': Decimal('10000.00')
            }
        ]
        
        # Expected tax calculations based on French flat tax formula
        # Formula: Taxable Gain = Withdrawal - (Acquisition Cost × (Withdrawal / Portfolio Value))
        # Note: Portfolio value is the value BEFORE withdrawal (passed to the function)
        self.expected_results = [
            {
                'acquisition_cost': Decimal('10000.00'),
                'taxable_gain': Decimal('0.00'),
                'cumulative_gains': Decimal('0.00')
            },
            {
                # Portfolio value = 8000 EUR (before withdrawal)
                # Acquisition portion = 10000 * (3000/8000) = 3750.00
                # Taxable gain = 3000 - 3750.00 = -750.00
                # New acquisition = 10000 - 3750.00 = 6250.00
                'acquisition_cost': Decimal('6250.00'),
                'taxable_gain': Decimal('-750.00'),
                'cumulative_gains': Decimal('-750.00')
            },
            {
                'acquisition_cost': Decimal('8250.00'),  # 6250 + 2000
                'taxable_gain': Decimal('0.00'),
                'cumulative_gains': Decimal('-750.00')
            },
            {
                # Portfolio value = 15000 EUR (before withdrawal)
                # Acquisition portion = 8250 * (5000/15000) = 2750.00
                # Taxable gain = 5000 - 2750.00 = 2250.00
                # New acquisition = 8250 - 2750.00 = 5500.00
                'acquisition_cost': Decimal('5500.00'),
                'taxable_gain': Decimal('2250.00'),
                'cumulative_gains': Decimal('1500.00')
            },
            {
                # Portfolio value = 10000 EUR (before withdrawal)
                # Acquisition portion = 5500 * (4000/10000) = 2200.00
                # Taxable gain = 4000 - 2200.00 = 1800.00
                # New acquisition = 5500 - 2200.00 = 3300.00
                'acquisition_cost': Decimal('3300.00'),
                'taxable_gain': Decimal('1800.00'),
                'cumulative_gains': Decimal('3300.00')
            }
        ]
    
    def tearDown(self):
        """Clean up test files"""
        if os.path.exists(self.test_dir):
            shutil.rmtree(self.test_dir)
        
        # Clean up rapports directory if created
        if os.path.exists('rapports'):
            for file in os.listdir('rapports'):
                if file.startswith('test_') or 'test' in file.lower():
                    os.remove(os.path.join('rapports', file))
    
    def test_end_to_end_workflow_with_mocked_apis(self):
        """Test complete workflow from API calls to Excel generation"""
        # Mock Binance API responses
        mock_operations = [
            FiatOperation(
                date=op['date'],
                operation_type=op['type'],
                amount_eur=op['amount'],
                timestamp=op['timestamp']
            )
            for op in self.test_operations
        ]
        
        with patch('config.config.Config.load_binance_keys') as mock_keys, \
             patch('clients.binance_client.Client') as mock_binance_sdk, \
             patch('clients.binance_client.BinanceClient.get_fiat_operations') as mock_get_ops, \
             patch('clients.binance_client.BinanceClient.get_portfolio_value_usd') as mock_get_portfolio, \
             patch('clients.frankfurter_client.FrankfurterClient.get_exchange_rate') as mock_get_rate, \
             patch('utils.logger.setup_logger') as mock_logger, \
             patch('utils.logger.get_logger') as mock_get_logger:
            
            # Configure mocks
            mock_keys.return_value = ('test_api_key', 'test_secret_key')
            mock_get_ops.return_value = mock_operations
            mock_logger.return_value = MagicMock()
            mock_get_logger.return_value = MagicMock()
            
            # Mock Binance SDK client
            mock_client_instance = MagicMock()
            mock_client_instance.get_account_status.return_value = {'success': True}
            mock_binance_sdk.return_value = mock_client_instance
            
            # Mock portfolio values for each operation
            def get_portfolio_side_effect(timestamp):
                for op in self.test_operations:
                    if op['timestamp'] == timestamp:
                        return op['portfolio_usd']
                return Decimal('0')
            
            mock_get_portfolio.side_effect = get_portfolio_side_effect
            
            # Mock exchange rates (all 0.92 for simplicity)
            mock_get_rate.return_value = Decimal('0.92')
            
            # Generate report
            output_path = os.path.join(self.test_dir, 'test_report.xlsx')
            
            # Manually run the workflow components
            binance_client = BinanceClient('test_key', 'test_secret')
            frankfurter_client = FrankfurterClient()
            tax_calculator = FlatTaxCalculator()
            portfolio_calculator = PortfolioValueCalculator()
            
            operations = binance_client.get_fiat_operations(self.test_year)
            
            report_rows = []
            for operation in operations:
                portfolio_value_usd = binance_client.get_portfolio_value_usd(operation.timestamp)
                operation_date = operation.date.date()
                exchange_rate = frankfurter_client.get_exchange_rate(operation_date, "USD", "EUR")
                
                portfolio_value_eur = None
                if operation.operation_type == "Retrait":
                    portfolio_value_eur = portfolio_calculator.convert_usd_to_eur(
                        portfolio_value_usd, exchange_rate
                    )
                
                if operation.operation_type == "Dépôt":
                    tax_calc = tax_calculator.process_deposit(operation.amount_eur)
                else:
                    tax_calc = tax_calculator.process_withdrawal(
                        operation.amount_eur, 
                        portfolio_value_eur
                    )
                
                report_row = TaxReportRow(
                    date=operation_date,
                    operation_type=operation.operation_type,
                    amount_eur=operation.amount_eur,
                    portfolio_value_usd=portfolio_value_usd,
                    exchange_rate=exchange_rate,
                    portfolio_value_eur=portfolio_value_eur,
                    acquisition_cost=tax_calc.acquisition_cost,
                    taxable_gain=tax_calc.taxable_gain,
                    cumulative_gains=tax_calc.cumulative_gains
                )
                
                report_rows.append(report_row)
            
            # Generate Excel report
            excel_writer = ExcelReportWriter()
            excel_path = excel_writer.create_report(report_rows, self.test_year, output_path)
            
            # Verify Excel file was created
            self.assertTrue(os.path.exists(excel_path))
            
            # Verify Excel content
            self._verify_excel_content(excel_path, report_rows)
    
    def test_excel_output_format(self):
        """Test that Excel output has correct format and calculations"""
        # Create test report rows
        report_rows = []
        tax_calculator = FlatTaxCalculator()
        
        for idx, op_data in enumerate(self.test_operations):
            if op_data['type'] == 'Dépôt':
                tax_calc = tax_calculator.process_deposit(op_data['amount'])
                portfolio_value_eur = None
            else:
                portfolio_value_eur = op_data['portfolio_eur']
                tax_calc = tax_calculator.process_withdrawal(
                    op_data['amount'],
                    portfolio_value_eur
                )
            
            report_row = TaxReportRow(
                date=op_data['date'].date(),
                operation_type=op_data['type'],
                amount_eur=op_data['amount'],
                portfolio_value_usd=op_data['portfolio_usd'],
                exchange_rate=op_data['exchange_rate'],
                portfolio_value_eur=portfolio_value_eur,
                acquisition_cost=tax_calc.acquisition_cost,
                taxable_gain=tax_calc.taxable_gain,
                cumulative_gains=tax_calc.cumulative_gains
            )
            
            report_rows.append(report_row)
        
        # Generate Excel
        output_path = os.path.join(self.test_dir, 'test_format.xlsx')
        excel_writer = ExcelReportWriter()
        excel_path = excel_writer.create_report(report_rows, self.test_year, output_path)
        
        # Load and verify
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Verify headers
        expected_headers = [
            "Date",
            "Type d'opération (Dépôt/Retrait Fiat)",
            "Montant en EUR",
            "Valeur portefeuille USD (après opération)",
            "Taux de change USD/EUR",
            "Valeur totale du portefeuille (EUR)",
            "Prix total d'acquisition restant (EUR)",
            "Plus-value imposable (EUR)",
            "Cumul plus-values (EUR)"
        ]
        
        for col_idx, expected_header in enumerate(expected_headers, start=1):
            actual_header = ws.cell(row=1, column=col_idx).value
            self.assertEqual(actual_header, expected_header)
        
        # Verify data rows
        for row_idx, report_row in enumerate(report_rows, start=2):
            # Date format
            date_cell = ws.cell(row=row_idx, column=1).value
            self.assertEqual(date_cell, report_row.date.strftime("%Y-%m-%d"))
            
            # Operation type
            op_type_cell = ws.cell(row=row_idx, column=2).value
            self.assertEqual(op_type_cell, report_row.operation_type)
            
            # Amount
            amount_cell = ws.cell(row=row_idx, column=3).value
            self.assertAlmostEqual(float(amount_cell), float(report_row.amount_eur), places=2)
            
            # Acquisition cost
            acq_cost_cell = ws.cell(row=row_idx, column=7).value
            self.assertAlmostEqual(float(acq_cost_cell), float(report_row.acquisition_cost), places=2)
            
            # Taxable gain
            gain_cell = ws.cell(row=row_idx, column=8).value
            self.assertAlmostEqual(float(gain_cell), float(report_row.taxable_gain), places=2)
            
            # Cumulative gains
            cumul_cell = ws.cell(row=row_idx, column=9).value
            self.assertAlmostEqual(float(cumul_cell), float(report_row.cumulative_gains), places=2)
        
        wb.close()
    
    def test_pdf_generation(self):
        """Test PDF report generation with --pdf flag"""
        # Create test report rows
        report_rows = []
        tax_calculator = FlatTaxCalculator()
        
        for op_data in self.test_operations[:3]:  # Use first 3 operations for speed
            if op_data['type'] == 'Dépôt':
                tax_calc = tax_calculator.process_deposit(op_data['amount'])
                portfolio_value_eur = None
            else:
                portfolio_value_eur = op_data['portfolio_eur']
                tax_calc = tax_calculator.process_withdrawal(
                    op_data['amount'],
                    portfolio_value_eur
                )
            
            report_row = TaxReportRow(
                date=op_data['date'].date(),
                operation_type=op_data['type'],
                amount_eur=op_data['amount'],
                portfolio_value_usd=op_data['portfolio_usd'],
                exchange_rate=op_data['exchange_rate'],
                portfolio_value_eur=portfolio_value_eur,
                acquisition_cost=tax_calc.acquisition_cost,
                taxable_gain=tax_calc.taxable_gain,
                cumulative_gains=tax_calc.cumulative_gains
            )
            
            report_rows.append(report_row)
        
        # Generate PDF
        output_path = os.path.join(self.test_dir, 'test_report.pdf')
        pdf_writer = PDFReportWriter()
        pdf_path = pdf_writer.create_report(report_rows, self.test_year, output_path)
        
        # Verify PDF file was created
        self.assertTrue(os.path.exists(pdf_path))
        
        # Verify file is not empty
        self.assertGreater(os.path.getsize(pdf_path), 0)
    
    def test_tax_calculations_match_expected(self):
        """Test that tax calculations match expected values from example"""
        tax_calculator = FlatTaxCalculator()
        
        for idx, op_data in enumerate(self.test_operations):
            if op_data['type'] == 'Dépôt':
                result = tax_calculator.process_deposit(op_data['amount'])
            else:
                result = tax_calculator.process_withdrawal(
                    op_data['amount'],
                    op_data['portfolio_eur']
                )
            
            expected = self.expected_results[idx]
            
            self.assertEqual(
                result.acquisition_cost,
                expected['acquisition_cost'],
                f"Operation {idx + 1}: Acquisition cost mismatch"
            )
            self.assertEqual(
                result.taxable_gain,
                expected['taxable_gain'],
                f"Operation {idx + 1}: Taxable gain mismatch"
            )
            self.assertEqual(
                result.cumulative_gains,
                expected['cumulative_gains'],
                f"Operation {idx + 1}: Cumulative gains mismatch"
            )
    
    def test_empty_operations_handling(self):
        """Test handling of year with no operations"""
        with patch('config.config.Config.load_binance_keys') as mock_keys, \
             patch('clients.binance_client.BinanceClient.get_fiat_operations') as mock_get_ops, \
             patch('utils.logger.setup_logger') as mock_logger, \
             patch('utils.logger.get_logger') as mock_get_logger:
            
            # Configure mocks
            mock_keys.return_value = ('test_api_key', 'test_secret_key')
            mock_get_ops.return_value = []  # No operations
            mock_logger.return_value = MagicMock()
            mock_get_logger.return_value = MagicMock()
            
            # Generate report with empty operations
            output_path = os.path.join(self.test_dir, 'test_empty.xlsx')
            excel_writer = ExcelReportWriter()
            excel_path = excel_writer.create_report([], self.test_year, output_path)
            
            # Verify file was created
            self.assertTrue(os.path.exists(excel_path))
            
            # Verify it has headers but no data rows
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Should have header row
            self.assertIsNotNone(ws.cell(row=1, column=1).value)
            
            # Should not have data in row 2
            self.assertIsNone(ws.cell(row=2, column=1).value)
            
            wb.close()
    
    def _verify_excel_content(self, excel_path: str, expected_rows: list):
        """Helper method to verify Excel content matches expected data"""
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Verify number of data rows
        # +1 for header, +2 for blank row and summary
        expected_row_count = len(expected_rows) + 1
        actual_data_rows = 0
        for row in ws.iter_rows(min_row=2):
            if row[0].value:  # Count rows with date values
                actual_data_rows += 1
            else:
                break
        
        self.assertEqual(actual_data_rows, len(expected_rows))
        
        # Verify each data row
        for idx, expected_row in enumerate(expected_rows, start=2):
            date_val = ws.cell(row=idx, column=1).value
            self.assertEqual(date_val, expected_row.date.strftime("%Y-%m-%d"))
            
            op_type_val = ws.cell(row=idx, column=2).value
            self.assertEqual(op_type_val, expected_row.operation_type)
            
            amount_val = ws.cell(row=idx, column=3).value
            self.assertAlmostEqual(float(amount_val), float(expected_row.amount_eur), places=2)
        
        wb.close()


if __name__ == '__main__':
    unittest.main()
