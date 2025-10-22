"""
Excel Report Writer for Binance Tax Reports.

This module generates formatted Excel files containing French tax calculations
for cryptocurrency operations.
"""

import os
from decimal import Decimal
from datetime import date, datetime
from dataclasses import dataclass
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from utils.logger import get_logger


class ExcelWriterError(Exception):
    """Exception raised for Excel writer errors."""
    pass


@dataclass
class TaxReportRow:
    """Complete row for Excel report"""
    date: date
    operation_type: str  # "Dépôt" or "Retrait"
    amount_eur: Decimal
    portfolio_value_usd: Decimal
    exchange_rate: Decimal
    portfolio_value_eur: Optional[Decimal]  # None for deposits
    acquisition_cost: Decimal
    taxable_gain: Decimal
    cumulative_gains: Decimal


class ExcelReportWriter:
    """Generates Excel tax report with proper formatting"""
    
    # Column headers in the exact order required
    COLUMN_HEADERS = [
        "Date",
        "Type d'opération (Dépôt/Retrait Fiat)",
        "Montant en EUR",
        "Valeur portefeuille USD (après opération)",
        "Taux de change USD/EUR",
        "Valeur totale du portefeuille (EUR)",
        "Prix total d'acquisition restant (EUR)",
        "Plus-value imposable (EUR)",
        "Cumul plus-values (EUR)"
    ]
    
    def __init__(self):
        """Initialize the Excel report writer"""
        self.logger = get_logger()
    
    def create_report(self, operations: List[TaxReportRow], year: int, 
                     output_path: Optional[str] = None) -> str:
        """
        Create Excel report file.
        
        Args:
            operations: List of processed operations with tax calculations
            year: Fiscal year
            output_path: Path for output file (optional, defaults to standard name)
            
        Returns:
            Path to the created file
            
        Raises:
            ExcelWriterError: If file creation fails
        """
        try:
            # Generate default filename if not provided
            if output_path is None:
                # Create rapports directory if it doesn't exist
                try:
                    os.makedirs("rapports", exist_ok=True)
                except OSError as e:
                    self.logger.error(f"Failed to create rapports directory: {e}")
                    raise ExcelWriterError(f"Failed to create output directory: {e}")
                
                output_path = f"rapports/Declaration_Fiscale_Crypto_{year}.xlsx"
            else:
                # Ensure the directory exists for custom paths
                output_dir = os.path.dirname(output_path)
                if output_dir:
                    try:
                        os.makedirs(output_dir, exist_ok=True)
                    except OSError as e:
                        self.logger.error(f"Failed to create output directory '{output_dir}': {e}")
                        raise ExcelWriterError(f"Failed to create output directory: {e}")
            
            # Check if file is writable (not locked by another process)
            if os.path.exists(output_path):
                try:
                    # Try to open the file to check if it's locked
                    with open(output_path, 'a'):
                        pass
                except IOError:
                    self.logger.error(f"Output file '{output_path}' is locked or not writable")
                    raise ExcelWriterError(
                        f"Cannot write to '{output_path}'. "
                        "The file may be open in another program. Please close it and try again."
                    )
            
            self.logger.info(f"Creating Excel report: {output_path}")
            
            # Create workbook and worksheet
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = f"Déclaration {year}"
            
            # Add headers
            self._add_headers(worksheet)
            
            # Add data rows
            self._add_data_rows(worksheet, operations)
            
            # Add summary row
            self._add_summary_row(worksheet, operations)
            
            # Format worksheet
            self._format_worksheet(worksheet)
            
            # Save workbook
            try:
                workbook.save(output_path)
                self.logger.info(f"Excel report saved successfully: {output_path}")
            except IOError as e:
                self.logger.error(f"Failed to save Excel file: {e}")
                raise ExcelWriterError(
                    f"Failed to save Excel file '{output_path}'. "
                    "Please check file permissions and ensure the file is not open in another program."
                )
            
            return output_path
            
        except ExcelWriterError:
            raise
        except Exception as e:
            self.logger.error(f"Unexpected error creating Excel report: {e}")
            raise ExcelWriterError(f"Unexpected error creating Excel report: {e}")

    def _add_headers(self, worksheet) -> None:
        """
        Add column headers to the worksheet.
        
        Args:
            worksheet: openpyxl worksheet object
        """
        for col_idx, header in enumerate(self.COLUMN_HEADERS, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    def _add_data_rows(self, worksheet, operations: List[TaxReportRow]) -> None:
        """
        Add data rows to the worksheet with formulas for calculated columns.
        
        Args:
            worksheet: openpyxl worksheet object
            operations: List of tax report rows
        """
        for row_idx, operation in enumerate(operations, start=2):
            # Date (YYYY-MM-DD format)
            worksheet.cell(row=row_idx, column=1).value = operation.date.strftime("%Y-%m-%d")
            
            # Operation type
            worksheet.cell(row=row_idx, column=2).value = operation.operation_type
            
            # Amount in EUR (2 decimal places)
            worksheet.cell(row=row_idx, column=3).value = float(operation.amount_eur)
            
            # Portfolio value USD (2 decimal places)
            worksheet.cell(row=row_idx, column=4).value = float(operation.portfolio_value_usd)
            
            # Exchange rate (4 decimal places for precision)
            worksheet.cell(row=row_idx, column=5).value = float(operation.exchange_rate)
            
            # Portfolio value EUR - FORMULA: =D{row}*E{row} (empty for deposits)
            if operation.operation_type == "Retrait":
                worksheet.cell(row=row_idx, column=6).value = f"=D{row_idx}*E{row_idx}"
            else:
                worksheet.cell(row=row_idx, column=6).value = ""
            
            # Acquisition cost - FORMULA based on previous row
            # For deposits: previous cost + deposit amount
            # For withdrawals: previous cost - (previous cost * (withdrawal / portfolio value))
            if row_idx == 2:
                # First row
                if operation.operation_type == "Dépôt":
                    worksheet.cell(row=row_idx, column=7).value = f"=C{row_idx}"
                else:  # Retrait
                    worksheet.cell(row=row_idx, column=7).value = f"=0-(0*(C{row_idx}/F{row_idx}))"
            else:
                # Subsequent rows
                if operation.operation_type == "Dépôt":
                    worksheet.cell(row=row_idx, column=7).value = f"=G{row_idx-1}+C{row_idx}"
                else:  # Retrait
                    worksheet.cell(row=row_idx, column=7).value = f"=G{row_idx-1}-(G{row_idx-1}*(C{row_idx}/F{row_idx}))"
            
            # Taxable gain - FORMULA: withdrawal - (acquisition cost portion)
            # For deposits: 0
            # For withdrawals: =C{row}-(previous_G * (C{row}/F{row}))
            if operation.operation_type == "Dépôt":
                worksheet.cell(row=row_idx, column=8).value = 0
            else:  # Retrait
                if row_idx == 2:
                    worksheet.cell(row=row_idx, column=8).value = f"=C{row_idx}-(0*(C{row_idx}/F{row_idx}))"
                else:
                    worksheet.cell(row=row_idx, column=8).value = f"=C{row_idx}-(G{row_idx-1}*(C{row_idx}/F{row_idx}))"
            
            # Cumulative gains - FORMULA: =I{prev_row}+H{row} (or just H{row} for first row)
            if row_idx == 2:
                worksheet.cell(row=row_idx, column=9).value = f"=H{row_idx}"
            else:
                worksheet.cell(row=row_idx, column=9).value = f"=I{row_idx-1}+H{row_idx}"
    
    def _add_summary_row(self, worksheet, operations: List[TaxReportRow]) -> None:
        """
        Add summary row with totals at the end of the worksheet.
        
        Args:
            worksheet: openpyxl worksheet object
            operations: List of tax report rows
        """
        if not operations:
            return
        
        # Calculate totals
        total_deposits = sum(
            op.amount_eur for op in operations if op.operation_type == "Dépôt"
        )
        total_withdrawals = sum(
            op.amount_eur for op in operations if op.operation_type == "Retrait"
        )
        total_taxable_gains = operations[-1].cumulative_gains if operations else Decimal("0")
        
        # Summary row starts after all data rows + 1 blank row
        summary_row = len(operations) + 3
        
        # Add "TOTAL" label
        cell = worksheet.cell(row=summary_row, column=1)
        cell.value = "TOTAL"
        cell.font = Font(bold=True, size=11)
        
        # Add total deposits
        cell = worksheet.cell(row=summary_row, column=2)
        cell.value = f"Dépôts: {float(total_deposits):.2f} EUR"
        cell.font = Font(bold=True)
        
        # Add total withdrawals
        cell = worksheet.cell(row=summary_row, column=3)
        cell.value = f"Retraits: {float(total_withdrawals):.2f} EUR"
        cell.font = Font(bold=True)
        
        # Add total taxable gains
        cell = worksheet.cell(row=summary_row, column=8)
        cell.value = float(total_taxable_gains)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    def _format_worksheet(self, worksheet) -> None:
        """
        Apply formatting to the worksheet.
        
        Args:
            worksheet: openpyxl worksheet object
        """
        # Set column widths for better readability
        column_widths = {
            1: 12,   # Date
            2: 30,   # Operation type
            3: 15,   # Amount EUR
            4: 25,   # Portfolio value USD
            5: 20,   # Exchange rate
            6: 25,   # Portfolio value EUR
            7: 30,   # Acquisition cost
            8: 25,   # Taxable gain
            9: 20    # Cumulative gains
        }
        
        for col_idx, width in column_widths.items():
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = width
        
        # Format numeric columns
        # Exchange rate with 4 decimals for precision
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            row[4].number_format = '0.0000'  # Column E: Exchange rate
            row[4].alignment = Alignment(horizontal='right')
        
        # Other numeric columns with 2 decimal places
        numeric_columns_2dp = [3, 4, 6, 7, 8, 9]  # C, D, F, G, H, I
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for col_idx in numeric_columns_2dp:
                cell = row[col_idx - 1]
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal='right')
        
        # Center align date and operation type columns
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            row[0].alignment = Alignment(horizontal='center')  # Date
            row[1].alignment = Alignment(horizontal='left')    # Operation type
        
        # Freeze the header row
        worksheet.freeze_panes = 'A2'
